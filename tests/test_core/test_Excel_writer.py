import unittest
from openpyxl import load_workbook
from src.core.excel_writer import create_excel_file

class TestExcelWriter(unittest.TestCase):
    def test_create_excel_file(self):
        # Your input dictionary
        data = {
            0: {'Fa': ['CPSC 1301K', 'CPSC 2105', 'CPSC 3165', 'CPSC 4000'], 'Sp': ['CPSC 1302', 'CPSC 3121'],
                'Su': ['CPSC 2108']},
            1: {'Fa': ['CPSC 3175'], 'Sp': [], 'Su': []},
            2: {'Fa': ['CPSC 4175'], 'Sp': ['CPSC 4176'], 'Su': []}
        }

        filename = "test_excel_file"
        
        # Call the function to create the Excel file
        create_excel_file(data, filename)

        # Load the Excel file
        workbook = load_workbook(f"{filename}.xlsx")

        # Check if the sheets are created as expected
        sheet_names = workbook.sheetnames
        expected_sheet_names = ['Year 0', 'Year 1', 'Year 2']
        self.assertEqual(sheet_names, expected_sheet_names, "Sheet names do not match")

        # Check the content of the first sheet (Year 0)
        year_0_sheet = workbook['Year 0']
        # You can add more specific checks based on your expectations

        # Close the workbook
        workbook.close()

