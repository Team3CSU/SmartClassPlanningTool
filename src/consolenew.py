
import json


def process_class_schedule(input_file):
    with open(input_file, 'r') as file:
        classschedule_data = json.load(file)
    terms = set()
    for course in classschedule_data:
        for term in classschedule_data[course]:
            terms.add(term)
    return classschedule_data

# prerequesite graph

import json
import networkx as nx


def process_prerequisites(input_file):
    with open(input_file, 'r') as file:
        prerequisites_data = json.load(file)
    return create_prerequisite_graph(prerequisites_data)


def create_prerequisite_graph(prerequisites_data):
    graph = nx.DiGraph()  # Create a directed graph

    for course, prerequisites in prerequisites_data.items():
        for prerequisite in prerequisites:
            graph.add_edge(course, prerequisite)  # Add edges representing prerequisites

    return graph


#  recommendation.py

from networkx import DiGraph
import os


def has_prerequisites(course: str, graph: DiGraph):
    # if graph.has_node(course):
    for edge in graph.edges():
        if edge[0] == course:
            return True
    return False


def get_prerequisite(course: str, graph: DiGraph):
    for edge in graph.edges():
        if edge[0] == course:
            return edge[1]


def get_prerequisites(course: str, graph: DiGraph):
    courses = []
    for edge in graph.edges():
        if edge[0] == course:
            courses.append(edge[1])
    return courses


def cleanCourseCode(coursesRaw: list, Courseshedule):
    courses = []
    for course in coursesRaw:
        if "SELECT" in course:
            courses.append(course)
            continue
        course = course if "*" not in course else course[:course.find("*")]
        course = course.strip()
        course = " ".join(course.split());
        if course in Courseshedule:
            courses.append(course)
        else:
            print('{} is missing'.format(course))
    return courses


def newAcademicYr():
    semester_dict = {"Fa": list().copy(), "Sp": list().copy(), "Su": list().copy()}
    return semester_dict


def sorter(deps):
    if (not deps):
        return -1
    depth = 0
    for code, crsdepth in deps:
        depth = max(depth, crsdepth)
    return depth


def extractPreq(courses_to_take, course, graph, lvl=0):
    dependencies = []
    if has_prerequisites(course, graph):
        preqCourses = get_prerequisites(course, graph)
        for preq in preqCourses:
            if preq is not None:
                if preq not in courses_to_take:
                    courses_to_take.append(preq)
                dependencies.append((preq, lvl))
                if has_prerequisites(preq, graph):
                    dep_in_depth, crsestake_indpth = extractPreq(courses_to_take, preq, graph, lvl + 1)
                    dependencies.extend(dep_in_depth)
                    courses_to_take = crsestake_indpth
    return dependencies, courses_to_take


def generate_degree_plan(text: dict, graph: DiGraph, Courseshedule: dict = None,logsfolder=None):
    courses_to_take = []
    logfile  = open(f'{logsfolder}/extracted_logs.txt', 'a')
    courses_plan_dict = {}
    coursecount = 0
    # run a loop over all the keys of courses that have to be taken
    for key in text.keys():
        # extract courses for each key
        coursesRaw = text[key]
        counter = 0
        limit = 0
        courses = cleanCourseCode(coursesRaw, Courseshedule)
        if logsfolder:
            print("\narea : "+ key,file = logfile)
            print("\n\t\traw =>",coursesRaw,file = logfile)
            print("\t\tclean=>",courses,file = logfile)
        print("*******"*20)
        
        for course in courses:
            dependencies = []
            if "SELECT".lower() in course.lower():
                # if there is a selection criteria, set it as a limit
                limit = int(course.split(" ")[1].strip())
            if limit > 0 and counter == limit:
                # if the selection criteria has been satisfied break the inner loop
                break
            if "SELECT".lower() not in course.lower():
                coursecount += 1
                original_course = course
                dependencies, courses_to_take = extractPreq(courses_to_take, original_course, graph)
                courses_plan_dict[original_course] = dependencies
                if original_course not in courses_to_take:
                    courses_to_take.append(original_course)
                counter = counter + 1
        if logsfolder:
            print("\n","courses_to_take",file = logfile)
            print(courses_to_take,file = logfile)
    # return courses_to_take,courses_plan_dict
    Sems = {0: newAcademicYr(), 1: newAcademicYr()}

    if logsfolder:
        print("\n\nExtracting prereq",file = logfile)
        for course in sorted(courses_plan_dict, key=lambda x: sorter(courses_plan_dict[x]), reverse=True):
            print("\t\tMain course: ",course,file = logfile)
            for dep in courses_plan_dict[course]:
                print("\t\t","\t"*(dep[1]+1),dep,file = logfile)
        print("Started planning",file = logfile)
    coursesTaken, planned = AcadPlanner(courses_plan_dict, Courseshedule, graph, Sems, courses_to_take, after=(0, -1),logfile=logfile)
    coursesTaken_last = AcadPlannerLast(courses_plan_dict, Courseshedule, Sems, courses_to_take,coursesTaken=coursesTaken,logfile=logfile)
    coursesTaken_c = [c for c, t in coursesTaken]

    return Sems


def AcadPlanner(courses_plan_dict, Courseshedule, graph, Sems, courses_to_take=None, coursesTaken=None, plandep=0,
                after=(0, -1),logfile=None):
    if not coursesTaken:
        coursesTaken = list()
    planed = None
    for course in sorted(courses_plan_dict, key=lambda x: sorter(courses_plan_dict[x]), reverse=True):
        if logfile:
            print("\t"*plandep,"planning course :",course,file = logfile)
        plannedmax = after
        planlst = [after]
        dependencies = courses_plan_dict.get(course, extractPreq([course], course, graph)[0])
        # if no dep, place it
        if len(dependencies):
            # if dep constains last skip it
            for dep in sorted(dependencies, key=lambda x: x[1], reverse=True):
                if(dep[0]=="LAST"):
                    return coursesTaken, planed
            # iterate dependecies depth wise
            lvl = -1
            for dep in sorted(dependencies, key=lambda x: x[1], reverse=True):
                if lvl == -1:
                    lvl = dep[1]
                cc = [c for c, c_ in coursesTaken]
                if dep[0] in cc:
                    continue
                if lvl > dep[1]:
                    lvl = dep[1]
                    plannedmax = max(planlst)
                    planlst = [after]
                if logfile:
                    print("\t"*plandep,"planning course :",dep[0],dep[1],file = logfile)
                ctt, planed = AcadPlanner({dep[0]: extractPreq([dep[0]], dep[0], graph)[0]}, Courseshedule, graph, Sems,
                                          [], coursesTaken, plandep=dep[1], after=plannedmax,logfile=logfile)
                coursesTaken = ctt
                assert (planed is not None)
                planlst.append(planed)

        cc = [c for c, c_ in coursesTaken]
        if course not in cc:
            if logfile:
                print("\t"*plandep,"planning course :",course,f"after => {after}",file = logfile)
            coursesTaken.append((course, plannedmax))
            planed = PlaceCourse(course, Courseshedule, plandep, Sems, max(planlst))

    return coursesTaken, planed

def AcadPlannerLast(courses_plan_dict, Courseshedule, Sems, courses_to_take,coursesTaken=None,logfile=None):
    if not coursesTaken:
        coursesTaken = list()
    planed = None
    yr = max(Sems)
    sem = 0
    print(Sems)
    for s in Sems[yr]:
        if len(Sems[yr][s]):
            sem = s
    try:
        sem = SemCodedict[0][sem]
    except:
        sem = SemCodedict[1][sem]
    print(yr,sem)
    plannedmax = (yr,sem)
    for course in sorted(courses_plan_dict, key=lambda x: sorter(courses_plan_dict[x]), reverse=True):
        dependencies = courses_plan_dict.get(course, [])
        if len(dependencies) == 1:
            for dep in sorted(dependencies, key=lambda x: x[1], reverse=True):
                if(dep[0]=="LAST"):
                    print("placing last")
                cc = [c for c, c_ in coursesTaken]
                if course not in cc:
                    if logfile:
                        print("planning course :",course,f"after => {plannedmax}",file = logfile)
                    coursesTaken.append((course, plannedmax))
                    planed = PlaceCourse(course, Courseshedule, 0, Sems, plannedmax,noincr=True)



SemCodedict = [{"Fa": 0, "Sp": 1, "Su": 2}, {0: 'Fa', 1: 'Sp', 2: 'Su'}]


def PlaceCourse(course, CourseSchedule, isdependency, Sems, after,noincr=False):
    """
    course: course name
    CourseSchedule: dictionary mapping course names to a list of semesters in which they are available
    isdependency: boolean indicating whether the course is a dependency
    Sems: plan of the semesters assigned with courses, where planned course details are updated
    after: semester from which we are planning
    """
    planned = None
    earliest_available = None
    # Check if the course is a dependency
    # Find the earliest available semester for the dependency course
    earliest_available = getEarliestAvailableSem(course, CourseSchedule, after,noincr)
    year, sem = earliest_available
    if year not in Sems:
        Sems[year] = newAcademicYr()
    Sems[year][SemCodedict[1][sem]].append(course)
    planned = earliest_available

    return planned

def validate_curr_course(courses_taken,course):
    """
    validates prerequesites
    
    """
    for c in courses_taken:
        pass
    return True

def incAfter(after):
    year, semCode = after
    semCode += 1
    if semCode > 2:
        semCode %= 3
        year += 1
    return (year, semCode)


def getEarliestAvailableSem(course, CourseSchedule, after,noincr=False):
    if not noincr:
        after = incAfter(after)
    courseAvail = CourseSchedule[course]
    # ##print()
    SemCodedict = {"Fa": 0, "Sp": 1, "Su": 2}, {0: 'Fa', 1: 'Sp', 2: 'Su'}
    courseAvailCode = [
        SemCodedict[0][i] for i in courseAvail
    ]
    while True:
        year, semCode = after
        sem = SemCodedict[1][semCode]
        if sem not in courseAvail:
            after = incAfter(after)
        else:
            return after

# data parser 

#  parsing pdf
import re
from pdfminer.high_level import extract_text
import os


def parse_course_requirements_from_pdf(input_file):
    try:
        text = extract_text(input_file)
        return parse_course_requirements(text)
    except Exception as e:
        print(f"Error reading PDF file: {str(e)}")
        return []


def parse_course_requirements(text):
    course_requirements = []
    current_category = None

    lines = text.split('\n')
    for line in lines:
        if line.strip():
            if re.match(r'^AREA [A-Z]+:', line):
                current_category = line.strip()
            elif current_category:
                match = re.match(r'\d+ Class(?:es)? in (.+)', line)
                if match:
                    course_requirements.append((current_category, match.group(1)))

    # TODO Extract further get list of courses to be done
    # print(course_requirements)
    return course_requirements


def extract_courses_from_pdf(pdf_file_path):
    course_data = {}  # Dictionary to store course data
    area = None  # Variable to track the current area

    text = extract_text(pdf_file_path)

    lines = text.split('\n')

    for line in lines:
        # Regular expressions to match course information
        area_match = re.search(r"AREA [A-Z]:\s+(.*)", line)
        course_match = re.findall(r"\b[A-Z]+\s+\d+[*@]?\b", line)

        if area_match:
            area = area_match.group(1)
        elif course_match:
            if area is not None:
                if area not in course_data:
                    course_data[area] = []
                course_data[area].extend(course_match)

    return course_data


def extract_and_store_courses(pdf_file_path,logsfolder = None)->list:
    course_data = {}  # Dictionary to store course data
    current_area = None
    current_courses = []

    text = extract_text(pdf_file_path)
    count = 0
    for line in text.split('\n'):
        if re.search(r"^\s*AREA [A-Z]:?", line):
            count += 1
            # Extract the area name
            try:
                current_area = re.search(r"^\s*AREA [A-Z]:\s*(.+)", line).group(1).strip()
            except:
                try:
                    current_area = re.search(r"^\s*(AREA [A-Z])", line).group(1).strip()
                except:
                    current_area = f"Unamed Area : {count}"
            current_courses = []
        #  CPSC 1223 CYBR MATH=> [A-Z]+\s+\d+[*] 
        elif re.search(r"Still Needed:.*?([A-Z]+\s+\d+[A-Z]?[*]?[^@])", line):
            # Extract course information (e.g., "1 Class in CPSC 1302*")
            match = re.findall(r"([A-Z]+\s+\d+[A-Z]?[*]?[^@])", line)
            current_courses.extend(match)
        elif current_area and not re.search(r"AREA [A-Z]:", line) and line.strip():
            # Continue to collect course information for the current area
            match = re.findall(r"([A-Z]+\s+\d+[A-Z]?[*]?[^@])", line)
            current_courses.extend(match)
        elif not current_area and line.strip():
            # Handle cases where the area is not explicitly mentioned
            match = re.findall(r"([A-Z]+\s+\d+[A-Z]?[*]?)", line)
            current_courses.extend(match)

        if current_area:
            course_data[current_area] = current_courses
    # write to log
    if logsfolder:
        with open(f'{logsfolder}/extracted_logs.txt', 'a') as filehandle:
            for key, value in course_data.items():
                filehandle.write("\narea"+str(key))
                for course in value:
                    filehandle.write("\n\t\tcourse :"+str(course))
    return course_data

# excelwriter.py

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment


def create_excel_file(data, filename):
    # Create a new Excel workbook
    workbook = Workbook()

    # Iterate through the data and create year-wise tables with bold headings
    for year, semesters in data.items():
        worksheet = workbook.create_sheet(title=f'Year {year}')

        # Set the column widths for better readability
        worksheet.column_dimensions['A'].width = 15
        worksheet.column_dimensions['B'].width = 40

        # Write the "Path to Graduation" heading in the first five rows
        heading_cell = worksheet.cell(row=worksheet.max_row, column=1, value="Path to Graduation")
        heading_cell.font = Font(bold=True)
        heading_cell.alignment = Alignment(horizontal="center")
        for semester, courses in semesters.items():
            if courses:
                # Write the semester heading and format it as bold
                semester_full = {'Fa': 'Fall','Sp': 'Spring','Su': 'Summer'}
                semester_heading = f'Semester: {semester_full[semester]}'
                semester_heading_cell = worksheet.cell(row=worksheet.max_row + 1, column=1, value=semester_heading)
                semester_heading_cell.font = Font(bold=True)

                # Write the courses
                for course in courses:
                    worksheet.cell(row=worksheet.max_row + 1, column=2, value=course)

    # Remove the default sheet created by openpyxl
    workbook.remove(workbook.active)

    # Save the workbook to a file
    print(f"plan saved as {filename}.xlsx")
    workbook.save(f'{filename}.xlsx')
    return filename
    

import os

# Get the path of the currently running Python script
script_path = os.path.realpath(__file__)

# Get the parent folder path
parent_folder = os.path.dirname(script_path)
parent_folder = os.path.dirname(parent_folder)

# Create the "logs" folder if it doesn't exist
logs_folder = os.path.join(parent_folder, "logs")

if not os.path.exists(logs_folder):
    os.makedirs(logs_folder)
    print(f"Created 'logs' folder at {logs_folder}")
else:
    print(f"'logs' folder already exists at {logs_folder}")


def validate_files(file_paths):
    if "pdf" not in file_paths[0]:
        return False
    if "json" not in file_paths[1]:
        return False
    if "json" not in file_paths[2]:
        return False
    return True


file_labels = ["Degree Works:", "Prerequisite Graph:", "Class Schedule:"]
file_entries = []


def feature1():

    # Manually prompt for each file path instead of using argparse
    f1 = input('Enter the path to the first file [Degree Works] (default: "./Software Design Development Section V01 Fall Semester 2023 CO - 10102023 - 1004 AM/Sample Input2.pdf"): ')
    f2 = input('Enter the path to the second file [Prerequisite Graph] (default: "./Software Design Development Section V01 Fall Semester 2023 CO - 10102023 - 1004 AM/preReq3.json"): ')
    f3 = input('Enter the path to the third file [Class Schedule] (default: "./Software Design Development Section V01 Fall Semester 2023 CO - 10102023 - 1004 AM/courseSchedule.json"): ')
    output_file = input('Enter the path to the results file: ')

    # Provide default values if no input is given
    f1 = f1 or "./Software Design Development Section V01 Fall Semester 2023 CO - 10102023 - 1004 AM/Sample Input2.pdf"
    f2 = f2 or "./Software Design Development Section V01 Fall Semester 2023 CO - 10102023 - 1004 AM/preReq3.json"
    f3 = f3 or "./Software Design Development Section V01 Fall Semester 2023 CO - 10102023 - 1004 AM/courseSchedule.json"

    # Rest of your code remains the same
    file_paths = [f1, f2, f3]
    are_files_valid = validate_files(file_paths)

    if are_files_valid:
        text = extract_and_store_courses(file_paths[0],logsfolder=logs_folder)
        print("\n\nExtracted courses:")
        for k in text:
            print(f"{k}  :                 {text[k]}")
        print("\n\ngraph:")
        graph = process_prerequisites(file_paths[1])
        # schedule
        print("\n\nschedule:")
        schedule = process_class_schedule(file_paths[2])
        courses_to_take = generate_degree_plan(text, graph, schedule,logsfolder=logs_folder)
        create_excel_file(courses_to_take, output_file)
    else:
        print("Please provide valid input files.")

def main():
    while True:
        # Displaying the menu
        print("\nMenu:")
        print("1. Feature 1 (Schedule Planning)")
        print("2. Feature 2 (Schedule Validation)")
        print("3. Exit")
        choice = input("Enter your choice (1, 2, or 3): ")

        if choice == '1':
            feature1()
        elif choice == '2':
            feature2()
        elif choice == '3':
            print("Exiting the program.")
            break
        else:
            print("Invalid choice. Please enter 1, 2, or 3.")


def feature2():
    # New code for Feature 2
    preq_json = input('Enter the path to the prerequisite JSON file: (default: "./Software Design Development Section V01 Fall Semester 2023 CO - 10102023 - 1004 AM/preReq3.json"): ')
    excel_file = input('Enter the path to the Excel file: (default: "./Software Design Development Section V01 Fall Semester 2023 CO - 10102023 - 1004 AM/plan3.xlsx"):')
    # Process the input files as required for the new feature
    # ...
    f1 = preq_json or  "./Software Design Development Section V01 Fall Semester 2023 CO - 10102023 - 1004 AM/preReq3.json"
    f2 = excel_file or "./Software Design Development Section V01 Fall Semester 2023 CO - 10102023 - 1004 AM/plan3.xlsx"

    courses=[]
    graph = None
    try:
        courses = extract_subjects_from_excel(f2)
    except:
        print("Error in processing the input files[excel file].")
        return
    try:
        graph = process_prerequisites(f1)
    except:
        print("Error in processing the input files.[networkx]")
        return

    try:
        courses_taken = []
        for course in courses:
            if validate_curr_course(courses_taken,course):
                courses_taken.append(course)
                # print(f"Course '{course}' added to the list of courses taken.")
            else:
                print("courses taken",courses_taken)
                print("course pre req: ",get_prerequisites(course, graph))
                print(f"Course '{course}' cannot be added to the list of courses taken")
                print("Schedule invalid: Course has prerequisite issue.")
                break
        print("Course plan Schedule is Valid")
    except Exception as e:
        print("Error:", e)


def extract_subjects_from_excel(file_path):
    import pandas as pd
    all_sheets = pd.read_excel(file_path, sheet_name=None)
    subjects = []
    for sheet_name, df in all_sheets.items():
        print(df)
        if 'Unnamed: 1' in df.columns:
            # Filter out NaN values and rows containing 'Semester'
            print(f"Sheet Name: {sheet_name}")
            subjects.extend(df.loc[~df['Unnamed: 1'].fillna('').str.contains('Semester', na=False), 'Unnamed: 1'].tolist())

    fianlsubjects=[]
    for subject in subjects:
        # print(subject)
        if type(subject)==str:
            fianlsubjects.append(subject)
    # Clean the subject names by extracting only alphanumeric characters
    subjects = [re.sub(r'[^A-Z0-9\s]', '', subject) for subject in fianlsubjects]

    return subjects


main()