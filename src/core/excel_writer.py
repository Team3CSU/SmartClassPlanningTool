from openpyxl import Workbook
from openpyxl.styles import Font, Alignment


def create_excel_file(data, filename):
    # Create a new Excel workbook
    workbook = Workbook()

    # Iterate through the data and create year-wise tables with bold headings
    for year, semesters in data.items():
        worksheet = workbook.create_sheet(title=f'Year {year}')

        # Set the column widths for better readability
        worksheet.column_dimensions['A'].width = 15
        worksheet.column_dimensions['B'].width = 40

        # Write the "Path to Graduation" heading in the first five rows
        heading_cell = worksheet.cell(row=worksheet.max_row, column=1, value="Path to Graduation")
        heading_cell.font = Font(bold=True)
        heading_cell.alignment = Alignment(horizontal="center")
        for semester, courses in semesters.items():
            if courses:
                # Write the semester heading and format it as bold
                semester_full = {
                    'Fa': 'Fall',
                    'Sp': 'Spring',
                    'Su': 'Summer'
                }
                semester_heading = f'Semester: {semester_full[semester]}'
                semester_heading_cell = worksheet.cell(row=worksheet.max_row + 1, column=1, value=semester_heading)
                semester_heading_cell.font = Font(bold=True)

                # Write the courses
                for course in courses:
                    course_cell = worksheet.cell(row=worksheet.max_row + 1, column=2, value=course)

    # Remove the default sheet created by openpyxl
    workbook.remove(workbook.active)

    # Save the workbook to a file
    print(f"plan saved as {filename}.xlsx")
    workbook.save(f'{filename}.xlsx')
    return filename


if __name__ == "__main__":
    # Your input dictionary
    data = {
        0: {'Fa': ['CPSC 1301K', 'CPSC 2105', 'CPSC 3165', 'CPSC 4000'], 'Sp': ['CPSC 1302', 'CPSC 3121'],
            'Su': ['CPSC 2108']},
        1: {'Fa': ['CPSC 3175'], 'Sp': [], 'Su': []},
        2: {'Fa': ['CPSC 4175'], 'Sp': ['CPSC 4176'], 'Su': []}
    }

    create_excel_file(data, "file.xlsx")
